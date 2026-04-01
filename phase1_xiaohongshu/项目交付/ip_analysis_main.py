"""
泡泡玛特 IP 热度时间序列分析 — 主分析 + Excel输出
输出: ip_timeseries_YYYYMMDD_HHMMSS.xlsx (6个Sheet)
"""
import math
from datetime import datetime
from collections import defaultdict
from pathlib import Path

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ip_analysis_clean import (
    load_all, calc_heat, heat_level, parse_likes, BASE_DIR
)

# ─── 配置 ──────────────────────────────────

# IP 固定配色（Excel条件格式用）
IP_COLORS = {
    '泡泡玛特': 'D32F2F', 'Labubu': 'FF6F00', 'Dimoo': '1976D2',
    'Molly': 'E91E63', 'Skullpanda': '7B1FA2', 'Zsiga': '388E3C',
    '小甜豆': 'F57C00', '星星人': '0097A7', 'Hirono小野': '5D4037',
    'Pucky': '689F38', 'CryBaby': '455A64',
}

HEADER_FILL = PatternFill(fill_type='solid', fgColor='1F4E79')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    bottom=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN


def set_col_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def auto_filter(ws):
    ws.auto_filter.ref = ws.dimensions


# ─── 加载数据 ──────────────────────────────

print('='*50)
print('泡泡玛特 IP 热度时间序列分析')
print('='*50)

chart_df, posts_df, comments_df = load_all()

# 顶级IP列表（按chart_data帖子数排序）
TOP_IPS = chart_df['ip'].value_counts().index.tolist()

# ─── Sheet1: IP热度总览 ───────────────────

print('\n[1/6] IP热度总览...')

ip_overview = []
for ip in TOP_IPS:
    sub = chart_df[chart_df['ip'] == ip]
    likes_list = sub['likes'].tolist()
    heat = calc_heat(likes_list)

    # 评论维度（从DB）
    db_sub = comments_df[comments_df['ip'] == ip]
    comment_count = len(db_sub)
    avg_comment_likes = db_sub['comment_likes'].mean() if len(db_sub) > 0 else 0

    # 综合热度 v2
    comment_ratio = comment_count / (len(sub) + 1)
    heat_v2 = heat * (1 + min(comment_ratio, 50) * 0.01)

    ip_overview.append({
        'ip': ip,
        'post_count': len(sub),
        'avg_likes': round(sub['likes'].mean()),
        'max_likes': sub['likes'].max(),
        'total_likes': sub['likes'].sum(),
        'comment_count': comment_count,
        'avg_comment_likes': round(avg_comment_likes, 1),
        'heat_v1': round(heat, 1),
        'heat_v2': round(heat_v2, 1),
        'level': heat_level(heat_v2),
    })

ip_overview.sort(key=lambda x: x['heat_v2'], reverse=True)
for i, row in enumerate(ip_overview, 1):
    row['rank'] = i

# ─── Sheet2: 月度热度趋势（帖子维度）──────

print('[2/6] 月度热度趋势...')

# 按月 × IP 聚合 chart_data
monthly_ip = chart_df.groupby(['ym', 'ip']).agg(
    post_count=('likes', 'size'),
    avg_likes=('likes', 'mean'),
    max_likes=('likes', 'max'),
    total_likes=('likes', 'sum'),
).reset_index()

# 计算每月每IP热度
monthly_ip['heat'] = monthly_ip.apply(
    lambda r: calc_heat([r['avg_likes']] * r['post_count']), axis=1
)

all_months = sorted(chart_df['ym'].unique())

# 构建宽表：每月一行，每IP一列
heat_pivot = monthly_ip.pivot_table(
    index='ym', columns='ip', values='heat', fill_value=0
).reindex(all_months, fill_value=0)

posts_pivot = monthly_ip.pivot_table(
    index='ym', columns='ip', values='post_count', fill_value=0
).reindex(all_months, fill_value=0)

likes_pivot = monthly_ip.pivot_table(
    index='ym', columns='ip', values='avg_likes', fill_value=0
).reindex(all_months, fill_value=0)

# ─── Sheet3: 评论热度趋势（评论维度）──────

print('[3/6] 评论热度趋势...')

comment_monthly_ip = comments_df.groupby(['ym', 'ip']).agg(
    comment_count=('id', 'size'),
    avg_comment_likes=('comment_likes', 'mean'),
    hot_comments=('comment_likes', lambda x: (x >= 10).sum()),
).reset_index()

comment_months = sorted(comments_df['ym'].unique())

comment_pivot = comment_monthly_ip.pivot_table(
    index='ym', columns='ip', values='comment_count', fill_value=0
).reindex(comment_months, fill_value=0)

comment_likes_pivot = comment_monthly_ip.pivot_table(
    index='ym', columns='ip', values='avg_comment_likes', fill_value=0
).reindex(comment_months, fill_value=0)

# ─── Sheet4: 热度衰减曲线 ─────────────────

print('[4/6] 热度衰减曲线...')

# 合并评论与帖子日期
merged = comments_df.merge(
    posts_df[['id', 'post_date_final']].rename(columns={'id': 'post_id'}),
    on='post_id', how='left'
)
# 计算评论距帖子发布的天数
merged['post_dt'] = pd.to_datetime(merged['post_date_final'], errors='coerce')
merged['comment_dt'] = pd.to_datetime(merged['clean_date'], errors='coerce')
merged['days_after'] = (merged['comment_dt'] - merged['post_dt']).dt.days

# 只保留合理范围 (0-365天)
valid = merged[(merged['days_after'] >= 0) & (merged['days_after'] <= 365)].copy()

# 按IP分桶统计
bins = [0, 7, 14, 21, 30, 60, 90, 180, 366]
labels = ['第1周', '第2周', '第3周', '第4周', '2月', '3月', '4-6月', '6月+']
valid['period'] = pd.cut(valid['days_after'], bins=bins, labels=labels, right=False)

decay_data = valid.groupby(['ip', 'period'], observed=True).size().reset_index(name='count')
decay_pivot = decay_data.pivot_table(
    index='period', columns='ip', values='count', fill_value=0
)

# 累计百分比
decay_cumsum = decay_pivot.cumsum()
decay_pct = decay_cumsum.div(decay_cumsum.iloc[-1]).fillna(0) * 100

# ─── Sheet5: 地域分布 ──────────────────────

print('[5/6] 地域分布...')

# 清理location
loc_df = comments_df[comments_df['location'] != ''].copy()
loc_df['location'] = loc_df['location'].str.strip()

# 总体地域分布
loc_total = loc_df['location'].value_counts().head(30)

# IP × 地域 矩阵
loc_ip = loc_df.groupby(['location', 'ip']).size().reset_index(name='count')
loc_ip_pivot = loc_ip.pivot_table(
    index='location', columns='ip', values='count', fill_value=0
)
# 按总评论数排序
loc_ip_pivot['总计'] = loc_ip_pivot.sum(axis=1)
loc_ip_pivot = loc_ip_pivot.sort_values('总计', ascending=False).head(25)

# 各IP地域集中度 HHI
hhi_data = []
for ip in TOP_IPS:
    ip_loc = loc_df[loc_df['ip'] == ip]['location'].value_counts()
    if len(ip_loc) == 0:
        continue
    total = ip_loc.sum()
    shares = (ip_loc / total) ** 2
    hhi = shares.sum()
    hhi_data.append({
        'ip': ip,
        'distinct_regions': len(ip_loc),
        'hhi': round(hhi, 4),
        'top1_region': ip_loc.index[0],
        'top1_share': round(ip_loc.iloc[0] / total * 100, 1),
    })

# ─── Sheet6: 帖子明细 ──────────────────────

print('[6/6] 帖子明细...')

post_detail = posts_df.sort_values('likes', ascending=False).copy()

# ─── 写Excel ──────────────────────────────

timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
output_path = BASE_DIR / f'ip_timeseries_{timestamp}.xlsx'

wb = openpyxl.Workbook()

# ── Sheet1: IP热度总览 ──
ws1 = wb.active
ws1.title = 'IP热度总览'
h1 = ['排名', 'IP', '帖子数', '平均点赞', '最高点赞', '总点赞',
      '深挖评论数', '均评论赞', '热度v1', '综合热度v2', '等级']
ws1.append(h1)
style_header(ws1, 1, len(h1))
for row in ip_overview:
    ws1.append([row['rank'], row['ip'], row['post_count'],
                row['avg_likes'], row['max_likes'], row['total_likes'],
                row['comment_count'], row['avg_comment_likes'],
                row['heat_v1'], row['heat_v2'], row['level']])
set_col_width(ws1, [6, 12, 8, 10, 10, 12, 10, 10, 10, 12, 8])
auto_filter(ws1)

# ── Sheet2: 月度热度趋势 ──
ws2 = wb.create_sheet('月度热度趋势')
# 写三块：发帖量 / 均赞 / 热度指数
blocks = [
    ('月度发帖量', posts_pivot),
    ('月度平均赞', likes_pivot.round(0).astype(int)),
    ('月度热度指数', heat_pivot.round(1)),
]
current_row = 1
for block_name, pivot in blocks:
    ws2.cell(row=current_row, column=1, value=block_name).font = Font(bold=True, size=12)
    current_row += 1
    # 表头
    ws2.cell(row=current_row, column=1, value='年月')
    for j, ip in enumerate(pivot.columns, 2):
        ws2.cell(row=current_row, column=j, value=ip)
    style_header(ws2, current_row, len(pivot.columns) + 1)
    current_row += 1
    # 数据
    for ym in pivot.index:
        ws2.cell(row=current_row, column=1, value=ym)
        for j, ip in enumerate(pivot.columns, 2):
            ws2.cell(row=current_row, column=j, value=pivot.loc[ym, ip])
        current_row += 1
    current_row += 2  # 空行

set_col_width(ws2, [10] + [12] * 11)

# ── Sheet3: 评论热度趋势 ──
ws3 = wb.create_sheet('评论热度趋势')
blocks3 = [
    ('月度评论量（按IP）', comment_pivot),
    ('月度均评论赞（按IP）', comment_likes_pivot.round(1)),
]
current_row = 1
for block_name, pivot in blocks3:
    ws3.cell(row=current_row, column=1, value=block_name).font = Font(bold=True, size=12)
    current_row += 1
    ws3.cell(row=current_row, column=1, value='年月')
    for j, ip in enumerate(pivot.columns, 2):
        ws3.cell(row=current_row, column=j, value=ip)
    style_header(ws3, current_row, len(pivot.columns) + 1)
    current_row += 1
    for ym in pivot.index:
        ws3.cell(row=current_row, column=1, value=ym)
        for j, ip in enumerate(pivot.columns, 2):
            ws3.cell(row=current_row, column=j, value=pivot.loc[ym, ip])
        current_row += 1
    current_row += 2

set_col_width(ws3, [10] + [12] * 8)

# ── Sheet4: 热度衰减曲线 ──
ws4 = wb.create_sheet('热度衰减曲线')
ws4.cell(row=1, column=1, value='评论时效分布（累计百分比）').font = Font(bold=True, size=12)
# 绝对数
ws4.cell(row=3, column=1, value='时间段')
for j, ip in enumerate(decay_pivot.columns, 2):
    ws4.cell(row=3, column=j, value=ip)
style_header(ws4, 3, len(decay_pivot.columns) + 1)
for i, period in enumerate(decay_pivot.index, 4):
    ws4.cell(row=i, column=1, value=str(period))
    for j, ip in enumerate(decay_pivot.columns, 2):
        ws4.cell(row=i, column=j, value=int(decay_pivot.loc[period, ip]))

# 累计百分比
gap_row = 4 + len(decay_pivot) + 1
ws4.cell(row=gap_row, column=1, value='累计百分比(%)').font = Font(bold=True, size=12)
gap_row += 1
ws4.cell(row=gap_row, column=1, value='时间段')
for j, ip in enumerate(decay_pct.columns, 2):
    ws4.cell(row=gap_row, column=j, value=ip)
style_header(ws4, gap_row, len(decay_pct.columns) + 1)
for i, period in enumerate(decay_pct.index, gap_row + 1):
    ws4.cell(row=i, column=1, value=str(period))
    for j, ip in enumerate(decay_pct.columns, 2):
        ws4.cell(row=i, column=j, value=round(decay_pct.loc[period, ip], 1))

set_col_width(ws4, [10] + [12] * 8)

# ── Sheet5: 地域分布 ──
ws5 = wb.create_sheet('地域分布')
# 地域×IP矩阵
h5 = ['地区'] + [c for c in loc_ip_pivot.columns]
ws5.append(h5)
style_header(ws5, 1, len(h5))
for loc in loc_ip_pivot.index:
    row_data = [loc] + [int(loc_ip_pivot.loc[loc, c]) for c in loc_ip_pivot.columns]
    ws5.append(row_data)

# HHI 地域集中度
gap_row = len(loc_ip_pivot) + 4
ws5.cell(row=gap_row, column=1, value='IP地域集中度(HHI)').font = Font(bold=True, size=12)
gap_row += 1
hhi_headers = ['IP', '覆盖地区数', 'HHI指数', '最大地区', '最大地区占比(%)']
for j, h in enumerate(hhi_headers, 1):
    ws5.cell(row=gap_row, column=j, value=h)
style_header(ws5, gap_row, len(hhi_headers))
for d in sorted(hhi_data, key=lambda x: x['hhi']):
    gap_row += 1
    ws5.append([d['ip'], d['distinct_regions'], d['hhi'],
                d['top1_region'], d['top1_share']])

set_col_width(ws5, [12] + [12] * 12)
auto_filter(ws5)

# ── Sheet6: 帖子明细 ──
ws6 = wb.create_sheet('帖子明细')
h6 = ['排名', 'IP', '标题', '点赞数', '发布日期', '深挖评论数', 'note_id']
ws6.append(h6)
style_header(ws6, 1, len(h6))
for i, (_, p) in enumerate(post_detail.iterrows(), 1):
    ws6.append([i, p['ip'], p['title'], p['likes'],
                p['post_date_final'] or '', p['comment_count'], p['note_id'] or ''])
set_col_width(ws6, [6, 12, 50, 10, 12, 10, 26])
auto_filter(ws6)

wb.save(str(output_path))
print(f'\n✅ Excel已保存: {output_path.name}')

# ─── 生成图表 ──────────────────────────────

print('\n生成图表...')
from ip_analysis_charts import generate_all_charts
generate_all_charts(
    chart_df=chart_df,
    posts_df=posts_df,
    comments_df=comments_df,
    heat_pivot=heat_pivot,
    posts_pivot=posts_pivot,
    likes_pivot=likes_pivot,
    comment_pivot=comment_pivot,
    decay_pct=decay_pct,
    loc_ip_pivot=loc_ip_pivot,
    ip_overview=ip_overview,
    output_dir=BASE_DIR / 'charts',
)

print('\n✅ 分析完成！')
