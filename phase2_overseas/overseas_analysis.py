"""
海外另类数据分析报告 — Pop Mart 海外热度时序分析 v2
投研级归一化分析框架：参与率/评论密度/IP份额/质量分层/滚动均线

输出: overseas_report_YYYYMMDD_HHMMSS.xlsx (7 Sheet) + charts/ (7 PNG)
数据源: overseas_data.db
"""
import os
import sqlite3
from datetime import datetime

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── 配置 ──────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'overseas_data.db')

plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'STHeiti']
plt.rcParams['axes.unicode_minus'] = False
plt.rcParams['figure.dpi'] = 150

IP_COLORS = {
    'Labubu': '#FF6F00', 'Dimoo': '#1976D2',
    'Molly': '#E91E63', 'Skullpanda': '#7B1FA2',
    'Pop Mart': '#D32F2F', 'Other': '#999999',
}
PLATFORM_COLORS = {
    'Amazon': '#FF9900', 'TikTok': '#000000', 'Instagram': '#E1306C',
}
TIKTOK_SOURCE_IP = {
    'tag/labubu': 'Labubu', 'tag/labubu lisa': 'Labubu',
    'tag/dimoo': 'Dimoo', 'tag/molly popmart': 'Molly',
    'tag/skullpanda': 'Skullpanda',
    'tag/pop mart': 'Pop Mart', 'tag/popmart unboxing': 'Pop Mart',
    'user/popmartglobal': 'Pop Mart',
}
INSTAGRAM_ACCOUNT_IP = {
    'hashtag:labubu': 'Labubu', 'hashtag:labubuthemonsters': 'Labubu',
    'hashtag:popmartlabubu': 'Labubu',
    'hashtag:popmart': 'Pop Mart', 'hashtag:popmartglobal': 'Pop Mart',
    'popmart': 'Pop Mart',
}
TIKTOK_BRAND_AUTHORS = {'popmartglobal'}
INSTAGRAM_BRAND_ACCOUNTS = {'popmart'}
COMMENT_HIGH_LIKES = 10
COMMENT_MED_LIKES = 3
MA_WINDOW = 4
RATING_COLORS = {5: '#4CAF50', 4: '#8BC34A', 3: '#FFC107', 2: '#FF9800', 1: '#F44336'}

# Excel
HEADER_FILL = PatternFill(fill_type='solid', fgColor='1F4E79')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
ANOMALY_FILL = PatternFill(fill_type='solid', fgColor='FFFF00')


# ─── 数据加载（个体级） ─────────────────────────────

def load_amazon_reviews(conn):
    df = pd.read_sql_query(
        'SELECT asin, ip, review_date, rating, verified FROM amazon_review_dates', conn)
    df['date'] = pd.to_datetime(df['review_date'], errors='coerce')
    df = df.dropna(subset=['date'])
    df['week'] = df['date'].dt.to_period('W-MON').dt.start_time
    return df


def load_tiktok_videos(conn):
    df = pd.read_sql_query(
        'SELECT video_id, author, title, views, likes, comments_count, create_time, source '
        'FROM tiktok_videos', conn)
    df['views'] = pd.to_numeric(df['views'], errors='coerce').fillna(0).astype(int)
    df['likes'] = pd.to_numeric(df['likes'], errors='coerce').fillna(0).astype(int)
    df['comments_count'] = pd.to_numeric(df['comments_count'], errors='coerce').fillna(0).astype(int)
    df['date'] = pd.to_datetime(pd.to_numeric(df['create_time'], errors='coerce'), unit='s', errors='coerce')
    df = df.dropna(subset=['date'])
    df['week'] = df['date'].dt.to_period('W-MON').dt.start_time
    df['ip'] = df['source'].map(TIKTOK_SOURCE_IP).fillna('Other')
    df['er'] = (df['likes'] + df['comments_count']) / df['views'].clip(lower=1)
    df['is_brand'] = df['author'].isin(TIKTOK_BRAND_AUTHORS)
    return df


def load_tiktok_comments(conn):
    cdf = pd.read_sql_query(
        'SELECT c.comment_id, c.video_id, c.comment_date, c.likes, c.author_name, '
        'v.source FROM tiktok_comments c LEFT JOIN tiktok_videos v ON c.video_id = v.video_id',
        conn)
    cdf['date'] = pd.to_datetime(cdf['comment_date'], errors='coerce')
    cdf = cdf.dropna(subset=['date'])
    cdf['week'] = cdf['date'].dt.to_period('W-MON').dt.start_time
    cdf['likes'] = pd.to_numeric(cdf['likes'], errors='coerce').fillna(0).astype(int)
    cdf['ip'] = cdf['source'].map(TIKTOK_SOURCE_IP).fillna('Other')
    cdf['quality'] = np.where(cdf['likes'] >= COMMENT_HIGH_LIKES, 'high',
                     np.where(cdf['likes'] >= COMMENT_MED_LIKES, 'medium', 'low'))
    return cdf


def load_instagram_posts(conn):
    df = pd.read_sql_query(
        'SELECT shortcode, account, caption, likes, comments_count, post_date '
        'FROM instagram_posts', conn)
    df['date'] = pd.to_datetime(df['post_date'], errors='coerce')
    df = df.dropna(subset=['date'])
    df['week'] = df['date'].dt.to_period('W-MON').dt.start_time
    df['likes'] = pd.to_numeric(df['likes'], errors='coerce').fillna(0).astype(int)
    df['comments_count'] = pd.to_numeric(df['comments_count'], errors='coerce').fillna(0).astype(int)
    df['ip'] = df['account'].map(INSTAGRAM_ACCOUNT_IP).fillna('Other')
    df['is_brand'] = df['account'].isin(INSTAGRAM_BRAND_ACCOUNTS)
    return df


def load_instagram_comments(conn):
    cdf = pd.read_sql_query(
        'SELECT c.comment_id, c.shortcode, c.comment_date, c.likes, c.author_name, '
        'p.account FROM instagram_comments c LEFT JOIN instagram_posts p ON c.shortcode = p.shortcode',
        conn)
    cdf['date'] = pd.to_datetime(cdf['comment_date'], errors='coerce')
    cdf = cdf.dropna(subset=['date'])
    cdf['week'] = cdf['date'].dt.to_period('W-MON').dt.start_time
    cdf['likes'] = pd.to_numeric(cdf['likes'], errors='coerce').fillna(0).astype(int)
    cdf['ip'] = cdf['account'].map(INSTAGRAM_ACCOUNT_IP).fillna('Other')
    cdf['quality'] = np.where(cdf['likes'] >= COMMENT_HIGH_LIKES, 'high',
                     np.where(cdf['likes'] >= COMMENT_MED_LIKES, 'medium', 'low'))
    return cdf


def load_amazon_snapshots(conn):
    try:
        df = pd.read_sql_query(
            'SELECT asin, ip, price_usd, rating, reviews, bought_monthly, in_stock '
            'FROM amazon_snapshots ORDER BY scraped_at DESC', conn)
        # 每个asin只保留最新快照
        return df.drop_duplicates(subset='asin', keep='first')
    except Exception:
        return pd.DataFrame()


# ─── 分析计算 ──────────────────────────────────────

def compute_weekly_platform(tt_vids, tt_coms, ig_posts, ig_coms, amz_revs):
    """每平台每周的归一化指标"""
    rows = []

    # TikTok
    for week, grp in tt_vids.groupby('week'):
        n = len(grp)
        coms = tt_coms[tt_coms['week'] == week]
        rows.append({
            'week': week, 'platform': 'TikTok',
            'content_count': n,
            'total_views': grp['views'].sum(),
            'total_likes': grp['likes'].sum(),
            'comment_count': len(coms),
            'avg_likes': grp['likes'].mean(),
            'avg_views': grp['views'].mean(),
            'comments_per_content': len(coms) / n if n else 0,
            'avg_er': grp['er'].mean(),
        })

    # Instagram
    for week, grp in ig_posts.groupby('week'):
        n = len(grp)
        coms = ig_coms[ig_coms['week'] == week]
        rows.append({
            'week': week, 'platform': 'Instagram',
            'content_count': n,
            'total_views': 0,
            'total_likes': grp['likes'].sum(),
            'comment_count': len(coms),
            'avg_likes': grp['likes'].mean(),
            'avg_views': 0,
            'comments_per_content': len(coms) / n if n else 0,
            'avg_er': 0,
        })

    # Amazon
    for week, grp in amz_revs.groupby('week'):
        rows.append({
            'week': week, 'platform': 'Amazon',
            'content_count': len(grp),
            'total_views': 0, 'total_likes': 0,
            'comment_count': len(grp),  # reviews = content = comments
            'avg_likes': 0, 'avg_views': 0,
            'comments_per_content': 1,  # each review is one unit
            'avg_er': 0,
        })

    df = pd.DataFrame(rows).sort_values(['platform', 'week'])
    # 4-week MA
    for plat in df['platform'].unique():
        mask = df['platform'] == plat
        for col in ['comments_per_content', 'avg_likes', 'avg_er']:
            df.loc[mask, f'ma4_{col}'] = (
                df.loc[mask, col].rolling(MA_WINDOW, min_periods=2).mean().values)
    return df


def compute_weekly_ip(tt_vids, tt_coms, ig_posts, ig_coms, amz_revs):
    """每IP每周指标 + IP份额"""
    rows = []

    # TikTok by IP
    tt_weekly_total = tt_coms.groupby('week').size()
    for (week, ip), grp in tt_vids.groupby(['week', 'ip']):
        coms = tt_coms[(tt_coms['week'] == week) & (tt_coms['ip'] == ip)]
        total_coms = tt_weekly_total.get(week, 1)
        rows.append({
            'week': week, 'platform': 'TikTok', 'ip': ip,
            'content_count': len(grp),
            'comment_count': len(coms),
            'avg_likes': grp['likes'].mean(),
            'avg_views': grp['views'].mean(),
            'avg_er': grp['er'].mean(),
            'comments_per_content': len(coms) / len(grp) if len(grp) else 0,
            'ip_share': len(coms) / total_coms * 100 if total_coms else 0,
        })

    # Instagram by IP
    ig_weekly_total = ig_coms.groupby('week').size()
    for (week, ip), grp in ig_posts.groupby(['week', 'ip']):
        coms = ig_coms[(ig_coms['week'] == week) & (ig_coms['ip'] == ip)]
        total_coms = ig_weekly_total.get(week, 1)
        rows.append({
            'week': week, 'platform': 'Instagram', 'ip': ip,
            'content_count': len(grp),
            'comment_count': len(coms),
            'avg_likes': grp['likes'].mean(),
            'avg_views': 0, 'avg_er': 0,
            'comments_per_content': len(coms) / len(grp) if len(grp) else 0,
            'ip_share': len(coms) / total_coms * 100 if total_coms else 0,
        })

    # Amazon by IP
    for (week, ip), grp in amz_revs.groupby(['week', 'ip']):
        valid = grp.dropna(subset=['rating'])
        rows.append({
            'week': week, 'platform': 'Amazon', 'ip': ip,
            'content_count': len(grp),
            'comment_count': len(grp),
            'avg_likes': 0, 'avg_views': 0, 'avg_er': 0,
            'comments_per_content': 1,
            'ip_share': 0,
            'avg_rating': valid['rating'].mean() if len(valid) else None,
        })

    return pd.DataFrame(rows).sort_values(['platform', 'ip', 'week'])


def compute_engagement_dist(tt_vids, ig_posts):
    """分位数分析"""
    rows = []
    quantiles = [0.10, 0.25, 0.50, 0.75, 0.90, 0.99]
    q_names = ['P10', 'P25', 'P50', 'P75', 'P90', 'P99']

    for ip, grp in tt_vids.groupby('ip'):
        for metric in ['views', 'likes', 'er']:
            s = grp[metric].dropna()
            if len(s) < 5:
                continue
            row = {'platform': 'TikTok', 'ip': ip, 'metric': metric,
                   'count': len(s), 'mean': s.mean(), 'std': s.std()}
            for q, name in zip(quantiles, q_names):
                row[name] = s.quantile(q)
            rows.append(row)

    for ip, grp in ig_posts.groupby('ip'):
        for metric in ['likes', 'comments_count']:
            s = grp[metric].dropna()
            if len(s) < 5:
                continue
            row = {'platform': 'Instagram', 'ip': ip, 'metric': metric,
                   'count': len(s), 'mean': s.mean(), 'std': s.std()}
            for q, name in zip(quantiles, q_names):
                row[name] = s.quantile(q)
            rows.append(row)

    return pd.DataFrame(rows)


def compute_cross_platform_index(plat_metrics):
    """首有效周=100 归一化指数"""
    result = {}
    for plat in ['Amazon', 'TikTok', 'Instagram']:
        sub = plat_metrics[plat_metrics['platform'] == plat].sort_values('week')
        if plat == 'Amazon':
            series = sub.set_index('week')['content_count']
        else:
            series = sub.set_index('week')['comments_per_content']
        # 找首个有效周（content >= 2 或 Amazon 有数据）
        valid = series[series > 0]
        if len(valid) < 2:
            continue
        base = valid.iloc[0] if valid.iloc[0] > 0 else 1
        indexed = series / base * 100
        ma4 = indexed.rolling(MA_WINDOW, min_periods=2).mean()
        result[plat] = pd.DataFrame({
            f'{plat}_raw': series, f'{plat}_idx': indexed, f'{plat}_ma4': ma4})
    if not result:
        return pd.DataFrame()
    merged = pd.concat(result.values(), axis=1)
    return merged.sort_index()


def compute_amazon_rating_dist(amz_revs):
    """Amazon 各IP评分分布"""
    valid = amz_revs.dropna(subset=['rating']).copy()
    valid['rating'] = valid['rating'].astype(int)
    dist = valid.groupby(['ip', 'rating']).size().reset_index(name='count')
    totals = valid.groupby('ip').size().reset_index(name='total')
    dist = dist.merge(totals, on='ip')
    dist['pct'] = dist['count'] / dist['total'] * 100
    return dist


def compute_comment_quality(tt_coms, ig_coms):
    """评论质量分层"""
    rows = []
    for plat, cdf in [('TikTok', tt_coms), ('Instagram', ig_coms)]:
        for week, grp in cdf.groupby('week'):
            n = len(grp)
            high = (grp['quality'] == 'high').sum()
            med = (grp['quality'] == 'medium').sum()
            low = (grp['quality'] == 'low').sum()
            rows.append({
                'week': week, 'platform': plat, 'total': n,
                'high': high, 'high_pct': high / n * 100 if n else 0,
                'medium': med, 'med_pct': med / n * 100 if n else 0,
                'low': low, 'low_pct': low / n * 100 if n else 0,
            })
    return pd.DataFrame(rows).sort_values(['platform', 'week'])


def compute_brand_vs_ugc(tt_vids):
    """TikTok 品牌 vs UGC"""
    tt_vids = tt_vids.copy()
    tt_vids['type'] = np.where(tt_vids['is_brand'], 'Brand', 'UGC')
    return tt_vids.groupby('type').agg(
        video_count=('video_id', 'size'),
        avg_views=('views', 'mean'),
        median_views=('views', 'median'),
        avg_likes=('likes', 'mean'),
        avg_er=('er', 'mean'),
        avg_comments=('comments_count', 'mean'),
    ).reset_index()


def detect_anomalies(series, window=8, sigma=2):
    rolling_mean = series.rolling(window, min_periods=4).mean()
    rolling_std = series.rolling(window, min_periods=4).std()
    upper = rolling_mean + sigma * rolling_std
    return series[series > upper].index.tolist()


# ─── 图表 ──────────────────────────────────────────

def _sparse_xticks(ax, labels, max_ticks=20):
    n = len(labels)
    step = max(1, n // max_ticks)
    ax.set_xticks(range(0, n, step))
    ax.set_xticklabels([labels[i] for i in range(0, n, step)],
                       rotation=45, ha='right', fontsize=7)


def plot_tiktok_er(plat_ip, charts_dir):
    """图1: TikTok 周均参与率趋势（按IP）"""
    print('  图1: TikTok ER 趋势...')
    tt = plat_ip[(plat_ip['platform'] == 'TikTok') & (plat_ip['ip'] != 'Other')]
    if tt.empty:
        return

    fig, ax1 = plt.subplots(figsize=(14, 6))
    # 按IP画ER折线
    for ip in tt['ip'].unique():
        sub = tt[tt['ip'] == ip].sort_values('week')
        if len(sub) < 3:
            continue
        x = range(len(sub))
        color = IP_COLORS.get(ip, '#999')
        # MA4
        ma4 = sub['avg_er'].rolling(MA_WINDOW, min_periods=2).mean()
        ax1.plot(x, ma4 * 100, label=f'{ip} (MA4)', color=color, linewidth=2)
        ax1.scatter(x, sub['avg_er'] * 100, color=color, s=15, alpha=0.4)

    ax1.set_ylabel('参与率 ER (%)', fontsize=12)
    ax1.set_title('TikTok 参与率趋势（按IP, (likes+comments)/views）', fontsize=14, fontweight='bold')
    ax1.legend(fontsize=9, loc='upper left')
    ax1.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    fig.savefig(os.path.join(charts_dir, 'chart_tiktok_er_trend.png'))
    plt.close(fig)


def plot_comments_density(plat_metrics, charts_dir):
    """图2: 三平台评论密度趋势"""
    print('  图2: 评论密度趋势...')
    fig, ax = plt.subplots(figsize=(14, 6))

    for plat in ['TikTok', 'Instagram']:
        sub = plat_metrics[plat_metrics['platform'] == plat].sort_values('week')
        if len(sub) < 3:
            continue
        x = range(len(sub))
        weeks = [w.strftime('%m-%d') for w in sub['week']]
        color = PLATFORM_COLORS[plat]
        # MA4 实线
        ax.plot(x, sub['ma4_comments_per_content'], label=f'{plat} (MA4)',
                color=color, linewidth=2.5)
        # 原始值散点
        ax.scatter(x, sub['comments_per_content'], color=color, s=12, alpha=0.3)
        _sparse_xticks(ax, weeks)

    ax.set_ylabel('评论密度（评论数/内容数）', fontsize=12)
    ax.set_title('各平台评论密度趋势（每条内容平均评论量）', fontsize=14, fontweight='bold')
    ax.legend(fontsize=11)
    ax.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    fig.savefig(os.path.join(charts_dir, 'chart_comments_density.png'))
    plt.close(fig)


def plot_ip_share(tt_coms, charts_dir):
    """图3: TikTok IP评论份额堆叠面积图"""
    print('  图3: TikTok IP份额...')
    tt = tt_coms[tt_coms['ip'] != 'Other'].copy()
    if tt.empty:
        return

    # 按周×IP计算份额
    weekly_ip = tt.groupby(['week', 'ip']).size().reset_index(name='count')
    weekly_total = tt.groupby('week').size().reset_index(name='total')
    weekly_ip = weekly_ip.merge(weekly_total, on='week')
    weekly_ip['share'] = weekly_ip['count'] / weekly_ip['total'] * 100

    pivot = weekly_ip.pivot_table(index='week', columns='ip', values='share', fill_value=0)
    pivot = pivot.sort_index()

    # 只保留近 30 周有数据的
    pivot = pivot[pivot.sum(axis=1) > 0].tail(30)
    if len(pivot) < 3:
        return

    fig, ax = plt.subplots(figsize=(14, 6))
    ips = pivot.columns.tolist()
    colors = [IP_COLORS.get(ip, '#999') for ip in ips]
    weeks = [w.strftime('%m-%d') for w in pivot.index]

    ax.stackplot(range(len(pivot)), *[pivot[ip].values for ip in ips],
                 labels=ips, colors=colors, alpha=0.85)
    _sparse_xticks(ax, weeks)
    ax.set_ylabel('评论份额 (%)', fontsize=12)
    ax.set_ylim(0, 100)
    ax.set_title('TikTok 各IP评论份额（近30周）', fontsize=14, fontweight='bold')
    ax.legend(loc='upper left', fontsize=9)
    plt.tight_layout()
    fig.savefig(os.path.join(charts_dir, 'chart_ip_share.png'))
    plt.close(fig)


def plot_cross_platform_index(cross_idx, charts_dir):
    """图4: 跨平台归一化指数"""
    print('  图4: 跨平台指数...')
    if cross_idx.empty:
        return

    # 只保留近52周
    if len(cross_idx) > 52:
        cross_idx = cross_idx.tail(52)

    fig, ax = plt.subplots(figsize=(14, 6))
    weeks = [w.strftime('%m-%d') for w in cross_idx.index]

    for plat in ['Amazon', 'TikTok', 'Instagram']:
        ma4_col = f'{plat}_ma4'
        idx_col = f'{plat}_idx'
        if ma4_col not in cross_idx.columns:
            continue
        color = PLATFORM_COLORS[plat]
        x = range(len(cross_idx))
        ax.plot(x, cross_idx[ma4_col], label=f'{plat} (MA4)',
                color=color, linewidth=2.5)
        ax.scatter(x, cross_idx[idx_col], color=color, s=10, alpha=0.25)

    ax.axhline(y=100, color='grey', linewidth=0.8, linestyle='--', alpha=0.5)
    _sparse_xticks(ax, weeks)
    ax.set_ylabel('归一化指数（首有效周=100）', fontsize=12)
    ax.set_title('跨平台热度指数（归一化对比）', fontsize=14, fontweight='bold')
    ax.legend(fontsize=11)
    ax.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    fig.savefig(os.path.join(charts_dir, 'chart_cross_platform_index.png'))
    plt.close(fig)


def plot_amazon_rating(rating_dist, charts_dir):
    """图5: Amazon 各IP评分分布"""
    print('  图5: Amazon 评分分布...')
    if rating_dist.empty:
        return

    fig, ax = plt.subplots(figsize=(10, 6))
    ips = rating_dist['ip'].unique()
    ratings = sorted(rating_dist['rating'].unique())
    n_ips = len(ips)
    bar_width = 0.15
    x = np.arange(n_ips)

    for i, r in enumerate(ratings):
        sub = rating_dist[rating_dist['rating'] == r]
        vals = [sub[sub['ip'] == ip]['pct'].values[0] if len(sub[sub['ip'] == ip]) else 0
                for ip in ips]
        color = RATING_COLORS.get(r, '#999')
        ax.bar(x + i * bar_width, vals, bar_width, label=f'{r}★', color=color, alpha=0.85)

    ax.set_xticks(x + bar_width * (len(ratings) - 1) / 2)
    ax.set_xticklabels(ips, fontsize=11)
    ax.set_ylabel('占比 (%)', fontsize=12)
    ax.set_title('Amazon 各IP评分分布', fontsize=14, fontweight='bold')
    ax.legend(fontsize=10)
    ax.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    fig.savefig(os.path.join(charts_dir, 'chart_amazon_rating.png'))
    plt.close(fig)


def plot_brand_vs_ugc(brand_ugc, charts_dir):
    """图6: TikTok 品牌 vs UGC"""
    print('  图6: 品牌 vs UGC...')
    if brand_ugc.empty or len(brand_ugc) < 2:
        return

    fig, axes = plt.subplots(1, 4, figsize=(16, 5))
    metrics = [
        ('avg_views', '平均播放量'),
        ('avg_likes', '平均点赞'),
        ('avg_er', '平均参与率'),
        ('avg_comments', '平均评论数'),
    ]
    colors = {'Brand': '#D32F2F', 'UGC': '#333333'}

    for ax, (col, title) in zip(axes, metrics):
        vals = brand_ugc.set_index('type')[col]
        bars = ax.bar(vals.index, vals.values,
                      color=[colors.get(t, '#999') for t in vals.index], alpha=0.85)
        ax.set_title(title, fontsize=11, fontweight='bold')
        ax.grid(axis='y', alpha=0.3)
        for bar, v in zip(bars, vals.values):
            label = f'{v:.1%}' if col == 'avg_er' else f'{v:,.0f}'
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height(),
                    label, ha='center', va='bottom', fontsize=9)

    fig.suptitle('TikTok: Pop Mart官号 vs 用户生成内容', fontsize=14, fontweight='bold')
    plt.tight_layout()
    fig.savefig(os.path.join(charts_dir, 'chart_brand_vs_ugc.png'))
    plt.close(fig)


def plot_comment_quality(quality_df, charts_dir):
    """图7: 评论质量分层趋势"""
    print('  图7: 评论质量...')
    fig, axes = plt.subplots(1, 2, figsize=(16, 6))
    quality_colors = {'high': '#4CAF50', 'medium': '#FFC107', 'low': '#E0E0E0'}

    for ax, plat in zip(axes, ['TikTok', 'Instagram']):
        sub = quality_df[quality_df['platform'] == plat].sort_values('week')
        if len(sub) < 3:
            continue
        # 只保留近30周
        sub = sub.tail(30)
        x = range(len(sub))
        weeks = [w.strftime('%m-%d') for w in sub['week']]

        ax.stackplot(x,
                     sub['high_pct'].values,
                     sub['med_pct'].values,
                     sub['low_pct'].values,
                     labels=['高互动(≥10赞)', '中互动(3-9赞)', '低互动(<3赞)'],
                     colors=[quality_colors['high'], quality_colors['medium'], quality_colors['low']],
                     alpha=0.85)
        _sparse_xticks(ax, weeks, max_ticks=10)
        ax.set_ylim(0, 100)
        ax.set_ylabel('占比 (%)')
        ax.set_title(f'{plat} 评论质量分层', fontsize=12, fontweight='bold')
        ax.legend(loc='lower left', fontsize=8)

    plt.tight_layout()
    fig.savefig(os.path.join(charts_dir, 'chart_comment_quality.png'))
    plt.close(fig)


# ─── Excel 输出 ────────────────────────────────────

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN


def auto_col_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)


def _write_df(ws, df, start_row=1, fmt=None):
    """将 DataFrame 写入 worksheet"""
    # 表头
    for c, col in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=c, value=col)
    style_header(ws, start_row, len(df.columns))
    # 数据
    for r, (_, row) in enumerate(df.iterrows(), start_row + 1):
        for c, col in enumerate(df.columns, 1):
            val = row[col]
            if fmt and col in fmt:
                val = fmt[col](val)
            if isinstance(val, (np.integer, np.int64)):
                val = int(val)
            elif isinstance(val, (np.floating, np.float64)):
                val = round(float(val), 2)
            ws.cell(row=r, column=c, value=val)
    return start_row + len(df) + 1


def write_excel(data, output_path):
    """7-sheet Excel"""
    print('\n[Excel] 生成报告...')
    wb = openpyxl.Workbook()

    # ── Sheet 1: Overview ──
    ws = wb.active
    ws.title = 'Overview'
    overview_rows = []
    for plat in ['Amazon', 'TikTok', 'Instagram']:
        sub = data['plat_metrics'][data['plat_metrics']['platform'] == plat]
        if sub.empty:
            continue
        latest = sub.sort_values('week').tail(4)
        r = {
            '平台': plat,
            '总内容数': int(sub['content_count'].sum()),
            '总评论数': int(sub['comment_count'].sum()),
            '均评论密度': round(sub['comments_per_content'].mean(), 1),
            '近4周均密度': round(latest['comments_per_content'].mean(), 1),
        }
        if plat == 'TikTok':
            r['均参与率(ER)'] = f"{sub['avg_er'].mean():.1%}"
            r['近4周均ER'] = f"{latest['avg_er'].mean():.1%}"
        else:
            r['均参与率(ER)'] = '-'
            r['近4周均ER'] = '-'

        # 热评占比
        qual = data['quality']
        pq = qual[qual['platform'] == plat]
        if not pq.empty:
            r['热评占比(≥10赞)'] = f"{pq['high_pct'].mean():.1f}%"
        else:
            r['热评占比(≥10赞)'] = '-'
        overview_rows.append(r)

    ov_df = pd.DataFrame(overview_rows)
    _write_df(ws, ov_df)
    auto_col_width(ws)

    # ── Sheet 2: Amazon ──
    ws2 = wb.create_sheet('Amazon')
    amz_ip = data['ip_metrics'][data['ip_metrics']['platform'] == 'Amazon']
    if not amz_ip.empty:
        pivot = amz_ip.pivot_table(index='week', columns='ip', values='content_count',
                                   aggfunc='sum', fill_value=0)
        pivot['Total'] = pivot.sum(axis=1)
        pivot = pivot[pivot['Total'] > 0]
        pivot_df = pivot.reset_index()
        pivot_df['week'] = pivot_df['week'].dt.strftime('%Y-%m-%d')
        next_row = _write_df(ws2, pivot_df)

        # 评分分布
        next_row += 1
        ws2.cell(row=next_row, column=1, value='评分分布')
        ws2.cell(row=next_row, column=1).font = Font(bold=True, size=12)
        next_row += 1
        rd = data['rating_dist']
        if not rd.empty:
            rd_show = rd[['ip', 'rating', 'count', 'pct']].copy()
            rd_show['pct'] = rd_show['pct'].round(1).astype(str) + '%'
            _write_df(ws2, rd_show, start_row=next_row)
            next_row += len(rd_show) + 2

        # 快照
        snap = data.get('snapshots')
        if snap is not None and not snap.empty:
            ws2.cell(row=next_row, column=1, value='最新快照')
            ws2.cell(row=next_row, column=1).font = Font(bold=True, size=12)
            next_row += 1
            snap_show = snap[['asin', 'ip', 'price_usd', 'rating', 'reviews', 'bought_monthly']].copy()
            _write_df(ws2, snap_show, start_row=next_row)

    auto_col_width(ws2)

    # ── Sheet 3: TikTok ──
    ws3 = wb.create_sheet('TikTok')
    tt_ip = data['ip_metrics'][(data['ip_metrics']['platform'] == 'TikTok') &
                                (data['ip_metrics']['ip'] != 'Other')]
    if not tt_ip.empty:
        tt_show = tt_ip[['week', 'ip', 'content_count', 'avg_views', 'avg_likes',
                         'avg_er', 'comment_count', 'comments_per_content', 'ip_share']].copy()
        tt_show['week'] = tt_show['week'].dt.strftime('%Y-%m-%d')
        tt_show['avg_views'] = tt_show['avg_views'].round(0).astype(int)
        tt_show['avg_likes'] = tt_show['avg_likes'].round(0).astype(int)
        tt_show['avg_er'] = (tt_show['avg_er'] * 100).round(2).astype(str) + '%'
        tt_show['comments_per_content'] = tt_show['comments_per_content'].round(1)
        tt_show['ip_share'] = tt_show['ip_share'].round(1).astype(str) + '%'
        tt_show.columns = ['Week', 'IP', '视频数', '均播放', '均点赞', 'ER',
                           '评论数', '评论密度', 'IP份额']
        _write_df(ws3, tt_show)
    auto_col_width(ws3)

    # ── Sheet 4: Instagram ──
    ws4 = wb.create_sheet('Instagram')
    ig_ip = data['ip_metrics'][(data['ip_metrics']['platform'] == 'Instagram') &
                                (data['ip_metrics']['ip'] != 'Other')]
    if not ig_ip.empty:
        ig_show = ig_ip[['week', 'ip', 'content_count', 'avg_likes',
                         'comment_count', 'comments_per_content', 'ip_share']].copy()
        ig_show['week'] = ig_show['week'].dt.strftime('%Y-%m-%d')
        ig_show['avg_likes'] = ig_show['avg_likes'].round(0).astype(int)
        ig_show['comments_per_content'] = ig_show['comments_per_content'].round(1)
        ig_show['ip_share'] = ig_show['ip_share'].round(1).astype(str) + '%'
        ig_show.columns = ['Week', 'IP', '帖子数', '均点赞', '评论数', '评论密度', 'IP份额']
        _write_df(ws4, ig_show)
    auto_col_width(ws4)

    # ── Sheet 5: Engagement Distribution ──
    ws5 = wb.create_sheet('Engagement')
    eng = data['engagement_dist']
    if not eng.empty:
        eng_show = eng.copy()
        for col in ['mean', 'std', 'P10', 'P25', 'P50', 'P75', 'P90', 'P99']:
            if col in eng_show.columns:
                eng_show[col] = eng_show[col].round(2)
        _write_df(ws5, eng_show)
    auto_col_width(ws5)

    # ── Sheet 6: Cross Platform ──
    ws6 = wb.create_sheet('Cross Platform')
    ci = data['cross_idx']
    if not ci.empty:
        ci_show = ci.reset_index()
        ci_show.rename(columns={'index': 'Week'}, inplace=True)
        if 'Week' in ci_show.columns:
            ci_show['Week'] = ci_show['Week'].dt.strftime('%Y-%m-%d')
        for col in ci_show.columns:
            if col != 'Week':
                ci_show[col] = ci_show[col].round(1)
        _write_df(ws6, ci_show)
    auto_col_width(ws6)

    # ── Sheet 7: Quality ──
    ws7 = wb.create_sheet('Quality')
    # 品牌 vs UGC
    bu = data['brand_ugc']
    next_row = 1
    if not bu.empty:
        ws7.cell(row=1, column=1, value='TikTok 品牌 vs UGC')
        ws7.cell(row=1, column=1).font = Font(bold=True, size=12)
        bu_show = bu.copy()
        bu_show['avg_views'] = bu_show['avg_views'].round(0).astype(int)
        bu_show['avg_likes'] = bu_show['avg_likes'].round(0).astype(int)
        bu_show['avg_er'] = (bu_show['avg_er'] * 100).round(2).astype(str) + '%'
        bu_show['avg_comments'] = bu_show['avg_comments'].round(1)
        bu_show.columns = ['类型', '视频数', '均播放', '中位播放', '均点赞', '均ER', '均评论']
        next_row = _write_df(ws7, bu_show, start_row=2)
        next_row += 1

    # 评论质量汇总
    qual = data['quality']
    if not qual.empty:
        ws7.cell(row=next_row, column=1, value='评论质量分层')
        ws7.cell(row=next_row, column=1).font = Font(bold=True, size=12)
        next_row += 1
        # 按平台汇总
        q_summary = qual.groupby('platform').agg(
            total=('total', 'sum'), high=('high', 'sum'),
            medium=('medium', 'sum'), low=('low', 'sum')).reset_index()
        q_summary['high_pct'] = (q_summary['high'] / q_summary['total'] * 100).round(1).astype(str) + '%'
        q_summary['med_pct'] = (q_summary['medium'] / q_summary['total'] * 100).round(1).astype(str) + '%'
        q_summary['low_pct'] = (q_summary['low'] / q_summary['total'] * 100).round(1).astype(str) + '%'
        q_show = q_summary[['platform', 'total', 'high', 'high_pct', 'medium', 'med_pct', 'low', 'low_pct']]
        q_show.columns = ['平台', '总评论', '高互动', '高占比', '中互动', '中占比', '低互动', '低占比']
        _write_df(ws7, q_show, start_row=next_row)

    auto_col_width(ws7)

    wb.save(output_path)
    print(f'  Saved: {output_path}')


# ─── 主入口 ────────────────────────────────────────

def main():
    print('=' * 55)
    print('Pop Mart 海外热度分析报告 v2（归一化框架）')
    print('=' * 55)

    conn = sqlite3.connect(DB_PATH)
    try:
        # 1. 加载个体级数据
        print('\n[1/4] 加载数据...')
        amz = load_amazon_reviews(conn)
        tt_vids = load_tiktok_videos(conn)
        tt_coms = load_tiktok_comments(conn)
        ig_posts = load_instagram_posts(conn)
        ig_coms = load_instagram_comments(conn)
        snapshots = load_amazon_snapshots(conn)

        print(f'  Amazon: {len(amz)} 条评论, {amz["ip"].nunique()} IP')
        print(f'  TikTok: {len(tt_vids)} 视频 (ER均值={tt_vids["er"].mean():.1%}), {len(tt_coms)} 评论')
        print(f'  Instagram: {len(ig_posts)} 帖子, {len(ig_coms)} 评论')
        print(f'  Amazon快照: {len(snapshots)} ASIN')

        # 2. 分析计算
        print('\n[2/4] 归一化分析...')
        plat_metrics = compute_weekly_platform(tt_vids, tt_coms, ig_posts, ig_coms, amz)
        ip_metrics = compute_weekly_ip(tt_vids, tt_coms, ig_posts, ig_coms, amz)
        engagement_dist = compute_engagement_dist(tt_vids, ig_posts)
        cross_idx = compute_cross_platform_index(plat_metrics)
        rating_dist = compute_amazon_rating_dist(amz)
        quality = compute_comment_quality(tt_coms, ig_coms)
        brand_ugc = compute_brand_vs_ugc(tt_vids)

        # 打印关键指标
        for plat in ['TikTok', 'Instagram']:
            sub = plat_metrics[plat_metrics['platform'] == plat]
            if not sub.empty:
                avg_density = sub['comments_per_content'].mean()
                print(f'  {plat} 平均评论密度: {avg_density:.1f}')
        tt_high = quality[quality['platform'] == 'TikTok']
        if not tt_high.empty:
            print(f'  TikTok 热评占比: {tt_high["high_pct"].mean():.1f}%')
        if not brand_ugc.empty:
            brand = brand_ugc[brand_ugc['type'] == 'Brand']
            ugc = brand_ugc[brand_ugc['type'] == 'UGC']
            if not brand.empty and not ugc.empty:
                print(f'  Brand均播: {brand["avg_views"].iloc[0]:,.0f}, UGC均播: {ugc["avg_views"].iloc[0]:,.0f}')

        # 3. 图表
        print('\n[3/4] 生成图表...')
        charts_dir = os.path.join(BASE_DIR, 'charts')
        os.makedirs(charts_dir, exist_ok=True)

        plot_tiktok_er(ip_metrics, charts_dir)
        plot_comments_density(plat_metrics, charts_dir)
        plot_ip_share(tt_coms, charts_dir)
        plot_cross_platform_index(cross_idx, charts_dir)
        plot_amazon_rating(rating_dist, charts_dir)
        plot_brand_vs_ugc(brand_ugc, charts_dir)
        plot_comment_quality(quality, charts_dir)

        # 4. Excel
        print('\n[4/4] 生成 Excel...')
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        xlsx_path = os.path.join(BASE_DIR, f'overseas_report_{ts}.xlsx')
        write_excel({
            'plat_metrics': plat_metrics,
            'ip_metrics': ip_metrics,
            'engagement_dist': engagement_dist,
            'cross_idx': cross_idx,
            'rating_dist': rating_dist,
            'quality': quality,
            'brand_ugc': brand_ugc,
            'snapshots': snapshots,
        }, xlsx_path)

        print(f'\n{"=" * 55}')
        print(f'Done!')
        print(f'  Excel: {xlsx_path}')
        print(f'  Charts: {charts_dir}/')
        print(f'{"=" * 55}')

    finally:
        conn.close()


if __name__ == '__main__':
    main()
